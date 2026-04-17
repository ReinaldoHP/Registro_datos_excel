"""
==============================================================================
Autor: Reinaldo Hurtado
Proyecto: Auditor de Facturas Excel
Año: 2026
------------------------------------------------------------------------------
"sistema": "Control de Órdenes de Servicio"
"version": "2.0"
"desarrollador": "Reinaldo Hurtado"
==============================================================================
"""
import os
import re
import json
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import subprocess
import zipfile

# Evitar ventana de cmd
CREATE_NO_WINDOW = 0x08000000

# Configurar CustomTkinter de forma global
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class DuplicateSelector(ctk.CTkToplevel):
    """Interfaz gráfica para resolver colisiones de facturas duplicadas en el sistema."""
    def __init__(self, parent, factura_id, options):
        super().__init__(parent)
        self.title(f"Duplicados - Fac {factura_id}")
        self.geometry("750x400")
        self.result = None
        
        # Etiqueta informativa
        ctk.CTkLabel(
            self, 
            text=f"Se han detectado {len(options)} posibles ubicaciones.\nSelecciona la carpeta correcta para la auditoría de {factura_id}:",
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
        ).pack(pady=15)

        # Marco con lista tipo Scrollable
        frame = ctk.CTkScrollableFrame(self, fg_color="#1E293B", corner_radius=10)
        frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=5)
        
        for opt in options:
            btn = ctk.CTkButton(
                frame, 
                text=str(opt),
                font=("Consolas", 12),
                fg_color="#334155",
                hover_color="#475569",
                anchor="w",
                command=lambda o=opt: self.on_select(Path(o))
            )
            btn.pack(fill="x", padx=5, pady=3)
        
        self.transient(parent)
        self.grab_set()
        parent.wait_window(self)

    def on_select(self, path):
        self.result = path
        self.destroy()


class InvoiceAuditor(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Auditor Factura - Configuración de Reglas (v2.0)")
        self.geometry("1000x550")
        self.configure(fg_color="#0F172A") # Fondo principal super oscuro #0F172A (aprox a #111827)
        
        self.fills = {
            'VERDE': PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid"),
            'AZUL': PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid"),
            'ROJO': PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
            'AMARILLO': PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
            'NARANJA': PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid"),
            'MORADO': PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid") # Ciruela/Morado claro
        }
        self.font_roja = Font(color="FF0000")
        
        # Variables de las rutas y configuraciones
        self.xl_file_var = tk.StringVar()
        self.search_root_var = tk.StringVar()
        self.save_path_var = tk.StringVar()
        self.empresa_var = tk.StringVar(value="General")
        
        self.empresas_opciones = [
            "General",
            "ADRES",
            "COOSALUD / SANIDAD MILITAR",
            "GRUPO SOLIDARIA (Solidaria, Axa ARL, Bolivar SOAT, Zurich)",
            "SEGUROS BOLIVAR ARL",
            "GRUPO ESTADO (Estado, HDI, Mapfre, Sura SOAT, Colmena)",
            "AXA COLPATRIA SEGURES ESCOLARES / MUNDIAL SEGUROS",
            "AXA COLPATRIA SOAT / EQUIDAD SEGUROS",
            "LA PREVISORA",
            "POSITIVA ARL Y SEGUROS ESCOLARES",
            "SURA EPS / SURA ARL"
        ]

        self.config_file = Path(__file__).parent / "auditor_config.json"
        
        # Primero construir, luego cargar archivo (evita variables fantasma)
        self.build_ui()
        self.load_config()

    def build_ui(self):
        # Frame Principal Múltiple (Grid)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # ------------------- SIDEBAR ------------------- #
        self.sidebar = ctk.CTkFrame(self, width=180, corner_radius=0, fg_color="#1E293B")
        self.sidebar.grid(row=0, column=0, sticky="nswe")
        self.sidebar.grid_rowconfigure(5, weight=1)

        # Logo / Título sidebar
        title_lbl = ctk.CTkLabel(self.sidebar, text="Auditor Factura", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color="#14B8A6")
        title_lbl.grid(row=0, column=0, padx=20, pady=(30, 40))

        # Íconos simulados con texto unicode minimalista (igual que en tu imagen)
        nav_items = [
            ("⊞  Panel", None),
            ("⚙  Ajustes", None),
            ("🕒  Historial", None),
            ("📁  Archivos", self.open_saved_folder)
        ]
        for i, (item, cmd) in enumerate(nav_items):
            btn = ctk.CTkButton(self.sidebar, text=item, fg_color="transparent", text_color="#CBD5E1", 
                                hover_color="#334155", font=ctk.CTkFont(size=14), anchor="w",
                                command=cmd)
            btn.grid(row=i+1, column=0, padx=20, pady=10, sticky="ew")

        # Firma Reinaldo HP
        firma_texto = (
            "Autor: Reinaldo Hurtado\n"
            "Sistema: Control de Órdenes de Servicio\n"
            "Versión: 2.0\n"
            "Desarrollador: Reinaldo Hurtado\n"
            "Año: 2026"
        )
        lbl_firma = ctk.CTkLabel(self.sidebar, text=firma_texto, text_color="#475569", font=ctk.CTkFont(size=11, weight="bold"), justify="left")
        lbl_firma.grid(row=6, column=0, pady=20, padx=15, sticky="s")

        # ---------------- MAIN CONTENT ----------------- #
        self.main_content = ctk.CTkFrame(self, fg_color="transparent")
        self.main_content.grid(row=0, column=1, sticky="nswe", padx=40, pady=30)
        
        # Título Arriba
        lbl_header = ctk.CTkLabel(self.main_content, text="Auditor Factura - Configuración de Reglas (v2.0)", 
                                  font=ctk.CTkFont(size=14, weight="normal"), text_color="#94A3B8")
        lbl_header.pack(anchor="w", pady=(0, 20))

        # Función para simular las cajas oscuras como entradas
        def create_input_card(parent, label_text, var, cmd_exam, is_combo=False):
            card = ctk.CTkFrame(parent, fg_color="#1E293B", corner_radius=10, border_width=1, border_color="#334155")
            card.pack(fill="x", pady=8)
            
            # Header interior
            lbl = ctk.CTkLabel(card, text=label_text, text_color="#94A3B8", font=ctk.CTkFont(size=12))
            lbl.pack(anchor="w", padx=15, pady=(10, 0))
            
            # Zona interactiva internar flontante
            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(fill="x", padx=15, pady=(5, 15))
            
            if is_combo:
                # Icono de "set de reglas"
                lbl_icon = ctk.CTkLabel(inner, text="☷ Rule-sets", text_color="#64748B", font=ctk.CTkFont(size=12))
                lbl_icon.pack(side="right", padx=10)

                combo = ctk.CTkComboBox(inner, variable=var, values=self.empresas_opciones, 
                                        fg_color="#0F172A", border_color="#334155", button_color="#334155", text_color="#14B8A6",
                                        command=lambda e: self.save_config(), height=35)
                combo.pack(side="left", fill="x", expand=True)
            else:
                entry = ctk.CTkEntry(inner, textvariable=var, fg_color="#0F172A", border_color="#334155", height=35, text_color="#E2E8F0")
                entry.pack(side="left", fill="x", expand=True, padx=(0, 15))
                btn = ctk.CTkButton(inner, text="Examinar", command=cmd_exam, fg_color="transparent", 
                                    border_width=1, border_color="#10B981", text_color="#10B981", hover_color="#064E3B", width=80)
                btn.pack(side="right")
                
            return card

        # Dibujar las 4 cajas
        create_input_card(self.main_content, "0. Regla de Empresa a Auditar", self.empresa_var, None, is_combo=True)
        create_input_card(self.main_content, "1. Archivo/Carpeta Raíz (Excel)", self.xl_file_var, self.sel_excel)
        create_input_card(self.main_content, "2. Carpeta de Destino (Soportes)", self.search_root_var, self.sel_search)
        create_input_card(self.main_content, "3. Ruta de Guardado (Auditoría)", self.save_path_var, self.sel_save)

        # Botón Central Neón Verde Inferior
        self.btn_run = ctk.CTkButton(self.main_content, text="INICIAR PROCESO DE AUDITORÍA", 
                                     font=ctk.CTkFont(size=14, weight="bold"),
                                     fg_color="#059669", hover_color="#047857", text_color="white",
                                     height=45, corner_radius=8, command=self.audit_process)
        self.btn_run.pack(pady=(20, 0))

    def load_config(self):
        if self.config_file.exists():
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.xl_file_var.set(config.get('xl_file', ''))
                    self.search_root_var.set(config.get('search_root', ''))
                    self.save_path_var.set(config.get('save_path', ''))
                    self.empresa_var.set(config.get('empresa', 'General'))
            except: pass

    def save_config(self):
        try:
            config = {
                'xl_file': self.xl_file_var.get(),
                'search_root': self.search_root_var.get(),
                'save_path': self.save_path_var.get(),
                'empresa': self.empresa_var.get()
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4)
        except: pass

    def sel_excel(self):
        current = self.xl_file_var.get()
        init_dir = os.path.dirname(current) if current else None
        path = filedialog.askopenfilename(
            title="Selecciona el Listado Excel", 
            filetypes=[("Archivos Excel", "*.xlsx")], 
            initialdir=init_dir
        )
        if path: 
            self.xl_file_var.set(path)
            self.save_config()

    def sel_search(self):
        current = self.search_root_var.get()
        init_dir = current if current else None
        path = filedialog.askdirectory(
            title="Selecciona la Carpeta de Soportes", 
            initialdir=init_dir
        )
        if path: 
            self.search_root_var.set(path)
            self.save_config()

    def sel_save(self):
        current = self.save_path_var.get()
        init_dir = os.path.dirname(current) if current else None
        init_file = os.path.basename(current) if current else None
        path = filedialog.asksaveasfilename(
            title="Guardar como", 
            defaultextension=".xlsx", 
            filetypes=[("Archivos Excel", "*.xlsx")], 
            initialdir=init_dir,
            initialfile=init_file
        )
        if path: 
            self.save_path_var.set(path)
            self.save_config()

    def open_saved_folder(self):
        """Abre la ruta donde se guardó el último excel auditado"""
        save_path = self.save_path_var.get()
        if save_path:
            folder_path = os.path.dirname(save_path)
            if os.path.exists(folder_path):
                # Usar explorer directo como forma genérica en Windows
                os.startfile(folder_path)
            else:
                messagebox.showerror("Error", "La ruta de guardado configurada no existe.")
        else:
            messagebox.showinfo("Ruta Vacía", "Selecciona primero una ruta de guardado.")

    def _extract_id(self, text):
        if text is None or str(text).strip() == "" or str(text).strip().lower() == 'nan': 
            return ""
        val = str(text).strip()
        match = re.search(r'(\d+)$', val)
        if match:
            return match.group(1).lstrip('0') or '0'
        return val

    def build_native_index(self, base_path, ids_to_search):
        index = {fid: [] for fid in ids_to_search if fid}
        ids_list = [fid for fid in ids_to_search if fid]
        
        if not ids_list: return {}
        print(f"🚀 Generando mapeo directo para {len(ids_list)} facturas del Excel...")
        
        for root, dirs, files in os.walk(base_path):
            current_dir = Path(root)
            for d in dirs:
                d_upper = d.upper().strip()
                for fid in ids_list:
                    if fid in d_upper:
                        index[fid].append(current_dir / d)
                        
            for f in files:
                if f.lower().endswith('.zip'):
                    f_upper = f.upper().strip()
                    f_clean = Path(f).stem.upper().strip()
                    for fid in ids_list:
                        if fid in f_clean or fid in f_upper:
                            index[fid].append(current_dir / f)

        found_count = 0
        for fid in index:
            index[fid] = list(set(index[fid]))
            if index[fid]:
                found_count += 1
                if fid in ["2127662", "2127695", "2127758"]:
                    print(f"✅ DEBUG - Carpeta encontrada para {fid}: {index[fid]}")
        return index

    def audit_process(self):
        xl_file = self.xl_file_var.get()
        search_root = self.search_root_var.get()
        save_path = self.save_path_var.get()

        if not xl_file or not search_root or not save_path:
            messagebox.showwarning("Faltan Rutas", "Por favor, completa las 3 rutas seleccionando un archivo o carpeta en cada una.")
            return

        # CTk uses configure instead of config
        self.btn_run.configure(text="PROCESANDO...", state="disabled")
        self.update()

        try:
            wb = load_workbook(xl_file)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("Error de Lectura", f"No se pudo abrir el Excel:\n{e}")
            self.btn_run.configure(text="INICIAR PROCESO DE AUDITORÍA", state="normal")
            return

        col_idx = None
        col_total_idx = None
        for col_num in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_num).value
            if val:
                val_upper = str(val).upper()
                if 'SFANUMFAC' in val_upper:
                    col_idx = col_num
                if 'TOTAL_FACTURADO' in val_upper:
                    col_total_idx = col_num

        if not col_idx:
            messagebox.showerror("Error", "No se encontró la columna 'SFANUMFAC'.")
            self.btn_run.configure(text="INICIAR PROCESO DE AUDITORÍA", state="normal")
            return

        excel_ids = set()
        for i in range(2, ws.max_row + 1):
            val = ws.cell(row=i, column=col_idx).value
            if val is not None and str(val).strip() != "" and str(val).strip().lower() != 'nan':
                if not ws.row_dimensions[i].hidden:
                    fid = self._extract_id(val)
                    if fid: excel_ids.add(fid)
        
        dir_map = self.build_native_index(search_root, excel_ids)

        idx_obs = ws.max_column + 1
        ws.cell(row=1, column=idx_obs).value = "RESULTADO_AUDITORIA"

        for i in range(2, ws.max_row + 1):
            if ws.row_dimensions[i].hidden:
                continue
            
            val = ws.cell(row=i, column=col_idx).value
            if val is None or str(val).strip() == "" or str(val).strip().lower() == 'nan':
                continue
                
            fid = self._extract_id(val)
            if not fid: continue
            
            matches = dir_map.get(fid, [])
            
            final_path = None
            if matches:
                pendientes = [p for p in matches if ("PENDIENTE" in p.stem.upper() or any(c.isalpha() for c in p.stem))]
                exact_folders = [p for p in matches if p.stem.upper().strip() == fid]
                prefix_folders = [p for p in matches if (p.stem.upper().strip().startswith(f"{fid}_") or p.stem.upper().strip().startswith(f"{fid} "))]
                
                # Sort to prefer directories over files (zips) if both exist
                pendientes.sort(key=lambda x: not x.is_dir())
                exact_folders.sort(key=lambda x: not x.is_dir())
                prefix_folders.sort(key=lambda x: not x.is_dir())

                if pendientes:
                    final_path = pendientes[0]
                elif exact_folders:
                    final_path = exact_folders[0]
                elif len(prefix_folders) == 1:
                    final_path = prefix_folders[0]
                elif len([p for p in matches if p.is_dir()]) == 1:
                    final_path = [p for p in matches if p.is_dir()][0]
                elif len(matches) > 1:
                    sel_dialog = DuplicateSelector(self, fid, matches)
                    final_path = sel_dialog.result
                elif len(matches) == 1:
                    final_path = matches[0]
                        
            fill = self.fills['ROJO']
            msg = "NO CARPETA"

            if final_path:
                empresa = self.empresa_var.get()
                count = 0
                invalid_count = 0
                pdf_files_list = []
                xml_count = 0
                
                try:
                    if final_path.suffix.lower() == '.zip':
                        with zipfile.ZipFile(final_path, 'r') as zf:
                            pdf_files_list = [f for f in zf.namelist() if f.lower().endswith('.pdf') and not Path(f).name.upper().startswith('DLP_')]
                            xml_count = len([f for f in zf.namelist() if f.lower().endswith('.xml') or (Path(f).name.lower().startswith('ad') and not f.lower().endswith('.pdf'))])
                    else:
                        pdf_files_list = [f.name for f in final_path.glob("*.pdf") if not f.name.upper().startswith('DLP_')]
                        xml_count = len([f for f in final_path.iterdir() if f.is_file() and (f.suffix.lower() == '.xml' or (f.name.lower().startswith('ad') and f.suffix.lower() != '.pdf'))])
                    
                    count = len(pdf_files_list)
                    
                    if empresa == "General":
                        if len(fid) != 7:
                            invalid_count += 4  # Forza a invalidar si no son 7 dígitos
                        for fname in pdf_files_list:
                            name_stem = Path(fname).stem
                            if "__" in name_stem or fid not in name_stem:
                                invalid_count += 1
                except: pass

                stem_upper = final_path.stem.upper()
                has_letters = any(c.isalpha() for c in stem_upper)
                
                if has_letters and "SURA" not in empresa and "FAC_" not in stem_upper:
                    desc = stem_upper.replace(fid, "").replace("_", " ").replace("-", " ").strip()
                    msg = desc if desc else "PENDIENTE"
                    fill = self.fills['AZUL']
                elif invalid_count > 0 and empresa == "General":
                    msg = f"MAL SOPORTADO ({invalid_count}/4)"
                    fill = self.fills['NARANJA']
                else:
                    if empresa == "General":
                        if count >= 4:
                            msg = "SIN RADICAR"
                            fill = self.fills['VERDE']
                        else:
                            msg = f"FALTAN SOPORTES ({count}/4)"
                            fill = self.fills['AMARILLO']
                    elif empresa == "ADRES":
                        if count >= 3 and xml_count > 0:
                            msg = "SIN RADICAR (+XML)"
                            fill = self.fills['VERDE']
                        else:
                            msg = f"FALTAN SOPORTES ({count} PDFs, {xml_count} XML)"
                            fill = self.fills['AMARILLO']
                    elif empresa == "COOSALUD / SANIDAD MILITAR":
                        if count > 0:
                            msg = "SIN RADICAR"
                            fill = self.fills['VERDE']
                        else:
                            msg = "VACÍO / FALTAN SOPORTES"
                            fill = self.fills['AMARILLO']
                    elif "GRUPO SOLIDARIA" in empresa:
                        if count == 2:
                            msg = "SIN RADICAR (2 docs)"
                            fill = self.fills['VERDE']
                        else:
                            msg = f"FALTAN SOPORTES ({count}/2)"
                            fill = self.fills['AMARILLO']
                    elif "BOLIVAR ARL" in empresa:
                        has_fac = any("FAC_" in f.upper() for f in pdf_files_list)
                        if count > 0 and has_fac:
                            msg = "SIN RADICAR (FAC OK)"
                            fill = self.fills['VERDE']
                        elif count > 0:
                            msg = "FALTA PDF 'FAC_'"
                            fill = self.fills['AMARILLO']
                        else:
                            msg = "VACÍO"
                            fill = self.fills['AMARILLO']
                    elif "GRUPO ESTADO" in empresa:
                        if count == 1:
                            msg = "SIN RADICAR (1 solo doc)"
                            fill = self.fills['VERDE']
                        else:
                            msg = f"ERROR APILAMIENTO ({count})"
                            fill = self.fills['AMARILLO']
                    elif "ESCOLARES / MUNDIAL" in empresa:
                        if count >= 3:
                            msg = "SIN RADICAR (3 docs)"
                            fill = self.fills['VERDE']
                        else:
                            msg = f"FALTAN SOPORTES ({count}/3)"
                            fill = self.fills['AMARILLO']
                    elif "SOAT / EQUIDAD" in empresa:
                        if count >= 4:
                            msg = "SIN RADICAR (4 docs)"
                            fill = self.fills['VERDE']
                        else:
                            msg = f"FALTAN SOPORTES ({count}/4)"
                            fill = self.fills['AMARILLO']
                    elif empresa == "LA PREVISORA":
                        if count >= 5:
                            msg = "SIN RADICAR (5+ docs)"
                            fill = self.fills['VERDE']
                        else:
                            msg = f"FALTAN SOPORTES ({count}/6)"
                            fill = self.fills['AMARILLO']
                    elif "POSITIVA" in empresa:
                        if count >= 3:
                            msg = "SIN RADICAR (Positiva)"
                            fill = self.fills['VERDE']
                        else:
                            msg = f"FALTAN SOPORTES ({count})"
                            fill = self.fills['AMARILLO']
                    elif "SURA" in empresa:
                        nit = "800218979"
                        has_nit = any(nit in f for f in pdf_files_list) or (nit in final_path.name)
                        if count > 0:
                            if has_nit:
                                msg = "SIN RADICAR (SURA OK)"
                                fill = self.fills['VERDE']
                            else:
                                msg = "FORMATO SURA INCORRECTO"
                                fill = self.fills['NARANJA']
                        else:
                            msg = "VACÍO / FALTAN SOPORTES"
                            fill = self.fills['AMARILLO']
                    else:
                        msg = f"REVISAR ({count})"
                        fill = self.fills['AMARILLO']

            is_sobre_costo = False
            if col_total_idx:
                try:
                    val_tot = ws.cell(row=i, column=col_total_idx).value
                    if val_tot is not None and float(val_tot) > 5000000:
                        is_sobre_costo = True
                except ValueError:
                    pass

            if is_sobre_costo:
                msg = "SOBRE COSTO"
                fill = self.fills['MORADO']

            for c in range(1, idx_obs + 1):
                cell = ws.cell(row=i, column=c)
                cell.fill = fill
                if is_sobre_costo:
                    cell.font = self.font_roja
            ws.cell(row=i, column=idx_obs).value = msg

        if save_path:
            try:
                wb.save(save_path)
                messagebox.showinfo("Completado", "¡Proceso terminado!\nArchivo guardado exitosamente.")
            except Exception as e:
                messagebox.showerror("Error al Guardar", f"No se pudo guardar el archivo:\n{e}")
                
        self.btn_run.configure(text="INICIAR PROCESO DE AUDITORÍA", state="normal")

if __name__ == "__main__":
    app = InvoiceAuditor()
    app.mainloop()
